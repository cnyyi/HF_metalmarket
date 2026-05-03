# -*- coding: utf-8 -*-
import logging
from datetime import datetime
from io import BytesIO
from utils.database import DBConnection
from utils.format_utils import format_date, format_datetime

logger = logging.getLogger(__name__)

GARBAGE_FEE_DICT_CODE = 'garbage_fee'
GARBAGE_FEE_DICT_NAME = '垃圾费'
GARBAGE_FEE_DICT_TYPE = 'expense_item_income'


def _get_garbage_fee_expense_type_id(cursor):
    cursor.execute("""
        SELECT DictID FROM Sys_Dictionary
        WHERE DictType = ? AND DictCode = ? AND IsActive = 1
    """, (GARBAGE_FEE_DICT_TYPE, GARBAGE_FEE_DICT_CODE))
    row = cursor.fetchone()
    if row:
        return row.DictID

    cursor.execute("""
        SELECT DictID FROM Sys_Dictionary
        WHERE DictType = ? AND DictName = ? AND IsActive = 1
    """, (GARBAGE_FEE_DICT_TYPE, GARBAGE_FEE_DICT_NAME))
    row = cursor.fetchone()
    if row:
        return row.DictID

    return None


class GarbageFeeService:

    def get_list(self, page=1, per_page=10, year=None, business_type=None,
                 status=None, search=None):
        with DBConnection() as conn:
            cursor = conn.cursor()

            base_query = """
                SELECT gf.GarbageFeeID, gf.MerchantID, gf.Year, gf.BusinessType,
                       gf.RentalArea, gf.UnitPrice, gf.MinAmount,
                       gf.CalculatedFee, gf.FinalFee, gf.ReceivableID,
                       gf.Status, gf.Description,
                       gf.CreateBy, gf.CreateTime, gf.UpdateBy, gf.UpdateTime,
                       m.MerchantName,
                       u.RealName AS CreateUserName
                FROM GarbageFee gf
                LEFT JOIN Merchant m ON gf.MerchantID = m.MerchantID
                LEFT JOIN [User] u ON gf.CreateBy = u.UserID
            """
            count_query = """
                SELECT COUNT(*) FROM GarbageFee gf
                LEFT JOIN Merchant m ON gf.MerchantID = m.MerchantID
            """
            summary_query = """
                SELECT ISNULL(COUNT(*), 0),
                       ISNULL(SUM(gf.RentalArea), 0),
                       ISNULL(SUM(gf.FinalFee), 0)
                FROM GarbageFee gf
                LEFT JOIN Merchant m ON gf.MerchantID = m.MerchantID
            """

            conditions = []
            params = []

            if year:
                conditions.append("gf.Year = ?")
                params.append(int(year))

            if business_type:
                conditions.append("gf.BusinessType = ?")
                params.append(business_type)

            if status:
                conditions.append("gf.Status = ?")
                params.append(status)

            if search:
                conditions.append("(m.MerchantName LIKE ? OR gf.Description LIKE ?)")
                p = f'%{search}%'
                params.extend([p, p])

            where_clause = ""
            if conditions:
                where_clause = " WHERE " + " AND ".join(conditions)
                base_query += where_clause
                count_query += where_clause
                summary_query += where_clause

            cursor.execute(count_query, tuple(params))
            total_count = cursor.fetchone()[0]

            cursor.execute(summary_query, tuple(params))
            summary_row = cursor.fetchone()
            summary = {
                'total_count': summary_row[0],
                'total_area': float(summary_row[1]),
                'total_fee': float(summary_row[2]),
            }

            offset = (page - 1) * per_page
            base_query += " ORDER BY gf.Year DESC, gf.GarbageFeeID DESC OFFSET ? ROWS FETCH NEXT ? ROWS ONLY"
            params.extend([offset, per_page])

            cursor.execute(base_query, tuple(params))
            rows = cursor.fetchall()

            items = []
            for row in rows:
                create_time = row.CreateTime
                create_time_str = create_time.strftime('%Y-%m-%d %H:%M') if create_time and hasattr(create_time, 'strftime') else ''
                items.append({
                    'garbage_fee_id': row.GarbageFeeID,
                    'merchant_id': row.MerchantID,
                    'merchant_name': row.MerchantName or '',
                    'year': row.Year,
                    'business_type': row.BusinessType or '',
                    'rental_area': float(row.RentalArea) if row.RentalArea else 0,
                    'unit_price': float(row.UnitPrice) if row.UnitPrice else 0,
                    'min_amount': float(row.MinAmount) if row.MinAmount else 0,
                    'calculated_fee': float(row.CalculatedFee) if row.CalculatedFee else 0,
                    'final_fee': float(row.FinalFee) if row.FinalFee else 0,
                    'receivable_id': row.ReceivableID,
                    'status': row.Status or '',
                    'description': row.Description or '',
                    'create_user_name': row.CreateUserName or '',
                    'create_time': create_time_str,
                })

            total_pages = (total_count + per_page - 1) // per_page if total_count > 0 else 1

            return {
                'items': items,
                'total': total_count,
                'page': page,
                'per_page': per_page,
                'total_pages': total_pages,
                'summary': summary,
            }

    def get_detail(self, garbage_fee_id):
        with DBConnection() as conn:
            cursor = conn.cursor()

            cursor.execute("""
                SELECT gf.GarbageFeeID, gf.MerchantID, gf.Year, gf.BusinessType,
                       gf.RentalArea, gf.UnitPrice, gf.MinAmount,
                       gf.CalculatedFee, gf.FinalFee, gf.ReceivableID,
                       gf.Status, gf.Description,
                       gf.CreateBy, gf.CreateTime, gf.UpdateBy, gf.UpdateTime,
                       m.MerchantName,
                       u.RealName AS CreateUserName
                FROM GarbageFee gf
                LEFT JOIN Merchant m ON gf.MerchantID = m.MerchantID
                LEFT JOIN [User] u ON gf.CreateBy = u.UserID
                WHERE gf.GarbageFeeID = ?
            """, (garbage_fee_id,))
            row = cursor.fetchone()
            if not row:
                return None

            create_time = row.CreateTime
            create_time_str = create_time.strftime('%Y-%m-%d %H:%M') if create_time and hasattr(create_time, 'strftime') else ''
            update_time = row.UpdateTime
            update_time_str = update_time.strftime('%Y-%m-%d %H:%M') if update_time and hasattr(update_time, 'strftime') else ''

            receivable_info = None
            if row.ReceivableID:
                cursor.execute("""
                    SELECT ReceivableID, Amount, PaidAmount, RemainingAmount, Status
                    FROM Receivable WHERE ReceivableID = ? AND IsActive = 1
                """, (row.ReceivableID,))
                rev_row = cursor.fetchone()
                if rev_row:
                    receivable_info = {
                        'receivable_id': rev_row.ReceivableID,
                        'amount': float(rev_row.Amount),
                        'paid_amount': float(rev_row.PaidAmount),
                        'remaining_amount': float(rev_row.RemainingAmount),
                        'status': rev_row.Status,
                    }

            return {
                'garbage_fee_id': row.GarbageFeeID,
                'merchant_id': row.MerchantID,
                'merchant_name': row.MerchantName or '',
                'year': row.Year,
                'business_type': row.BusinessType or '',
                'rental_area': float(row.RentalArea) if row.RentalArea else 0,
                'unit_price': float(row.UnitPrice) if row.UnitPrice else 0,
                'min_amount': float(row.MinAmount) if row.MinAmount else 0,
                'calculated_fee': float(row.CalculatedFee) if row.CalculatedFee else 0,
                'final_fee': float(row.FinalFee) if row.FinalFee else 0,
                'receivable_id': row.ReceivableID,
                'receivable_info': receivable_info,
                'status': row.Status or '',
                'description': row.Description or '',
                'create_user_name': row.CreateUserName or '',
                'create_time': create_time_str,
                'update_time': update_time_str,
            }

    def get_preview(self, year):
        with DBConnection() as conn:
            cursor = conn.cursor()

            cursor.execute("""
                SELECT m.MerchantID, m.MerchantName, m.BusinessType,
                       ISNULL(SUM(cp.Area), 0) AS TotalArea
                FROM Merchant m
                INNER JOIN Contract c ON m.MerchantID = c.MerchantID
                INNER JOIN ContractPlot cp ON c.ContractID = cp.ContractID
                WHERE c.Status = N'生效'
                  AND c.StartDate <= ?
                  AND c.EndDate >= ?
                GROUP BY m.MerchantID, m.MerchantName, m.BusinessType
                ORDER BY m.MerchantName
            """, (f'{year}-12-31', f'{year}-01-01'))
            rows = cursor.fetchall()

            items = []
            skipped = []

            for row in rows:
                merchant_id = row.MerchantID
                merchant_name = row.MerchantName or ''
                business_type = row.BusinessType or ''
                total_area = float(row.TotalArea)

                cursor.execute("""
                    SELECT GarbageFeeID FROM GarbageFee
                    WHERE MerchantID = ? AND Year = ?
                """, (merchant_id, year))
                if cursor.fetchone():
                    skipped.append({
                        'merchant_id': merchant_id,
                        'merchant_name': merchant_name,
                        'reason': '该年度已存在垃圾费记录',
                    })
                    continue

                unit_price = 0
                min_amount = 0
                if business_type:
                    cursor.execute("""
                        SELECT UnitPrice, MinAmount FROM Sys_Dictionary
                        WHERE DictType = N'business_type' AND DictName = ? AND IsActive = 1
                    """, (business_type,))
                    dict_row = cursor.fetchone()
                    if dict_row:
                        unit_price = float(dict_row.UnitPrice) if dict_row.UnitPrice else 0
                        min_amount = float(dict_row.MinAmount) if dict_row.MinAmount else 0

                if not business_type or (unit_price == 0 and min_amount == 0):
                    skipped.append({
                        'merchant_id': merchant_id,
                        'merchant_name': merchant_name,
                        'reason': '未设置业态类型或字典中无对应单价/保底',
                    })
                    continue

                calculated_fee = unit_price * total_area
                final_fee = max(calculated_fee, min_amount)

                items.append({
                    'merchant_id': merchant_id,
                    'merchant_name': merchant_name,
                    'business_type': business_type,
                    'rental_area': total_area,
                    'unit_price': unit_price,
                    'min_amount': min_amount,
                    'calculated_fee': round(calculated_fee, 2),
                    'final_fee': round(final_fee, 2),
                })

            return {
                'items': items,
                'skipped': skipped,
                'total_count': len(items) + len(skipped),
                'generate_count': len(items),
                'skip_count': len(skipped),
            }

    def batch_generate(self, year, created_by=None):
        with DBConnection() as conn:
            cursor = conn.cursor()

            expense_type_id = _get_garbage_fee_expense_type_id(cursor)
            if not expense_type_id:
                raise ValueError("未找到垃圾费费用类型字典项，请先执行数据库迁移脚本")

            cursor.execute("""
                SELECT m.MerchantID, m.MerchantName, m.BusinessType,
                       ISNULL(SUM(cp.Area), 0) AS TotalArea
                FROM Merchant m
                INNER JOIN Contract c ON m.MerchantID = c.MerchantID
                INNER JOIN ContractPlot cp ON c.ContractID = cp.ContractID
                WHERE c.Status = N'生效'
                  AND c.StartDate <= ?
                  AND c.EndDate >= ?
                GROUP BY m.MerchantID, m.MerchantName, m.BusinessType
                ORDER BY m.MerchantName
            """, (f'{year}-12-31', f'{year}-01-01'))
            rows = cursor.fetchall()

            success_count = 0
            skip_count = 0
            errors = []

            for row in rows:
                merchant_id = row.MerchantID
                merchant_name = row.MerchantName or ''
                business_type = row.BusinessType or ''
                total_area = float(row.TotalArea)

                try:
                    cursor.execute("""
                        SELECT GarbageFeeID FROM GarbageFee
                        WHERE MerchantID = ? AND Year = ?
                    """, (merchant_id, year))
                    if cursor.fetchone():
                        skip_count += 1
                        continue

                    unit_price = 0
                    min_amount = 0
                    if business_type:
                        cursor.execute("""
                            SELECT UnitPrice, MinAmount FROM Sys_Dictionary
                            WHERE DictType = N'business_type' AND DictName = ? AND IsActive = 1
                        """, (business_type,))
                        dict_row = cursor.fetchone()
                        if dict_row:
                            unit_price = float(dict_row.UnitPrice) if dict_row.UnitPrice else 0
                            min_amount = float(dict_row.MinAmount) if dict_row.MinAmount else 0

                    if not business_type or (unit_price == 0 and min_amount == 0):
                        skip_count += 1
                        continue

                    calculated_fee = round(unit_price * total_area, 2)
                    final_fee = round(max(calculated_fee, min_amount), 2)

                    cursor.execute("""
                        INSERT INTO GarbageFee (MerchantID, Year, BusinessType, RentalArea,
                                                UnitPrice, MinAmount, CalculatedFee, FinalFee,
                                                Status, CreateBy, CreateTime)
                        OUTPUT INSERTED.GarbageFeeID
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, N'待收取', ?, GETDATE())
                    """, (merchant_id, year, business_type, total_area,
                          unit_price, min_amount, calculated_fee, final_fee,
                          created_by))
                    gf_row = cursor.fetchone()
                    garbage_fee_id = gf_row[0]

                    cursor.execute("""
                        INSERT INTO Receivable (MerchantID, ExpenseTypeID, Amount, DueDate,
                                                Status, PaidAmount, RemainingAmount,
                                                Description, CustomerType, CustomerID,
                                                ReferenceType, ReferenceID)
                        OUTPUT INSERTED.ReceivableID
                        VALUES (?, ?, ?, ?, N'未付款', 0, ?, N'Merchant', ?, N'garbage_fee', ?)
                    """, (merchant_id, expense_type_id, final_fee,
                          f'{year}-12-31',
                          final_fee,
                          merchant_id, garbage_fee_id))
                    rev_row = cursor.fetchone()
                    receivable_id = rev_row[0]

                    cursor.execute("""
                        UPDATE GarbageFee SET ReceivableID = ? WHERE GarbageFeeID = ?
                    """, (receivable_id, garbage_fee_id))

                    success_count += 1

                except Exception as e:
                    errors.append(f'{merchant_name}: {str(e)}')
                    logger.error(f"批量生成垃圾费失败: MerchantID={merchant_id}, {e}")

            conn.commit()

            logger.info(f"垃圾费批量生成完成: 年度={year}, 成功={success_count}, 跳过={skip_count}, 错误={len(errors)}")
            return {
                'success_count': success_count,
                'skip_count': skip_count,
                'error_count': len(errors),
                'errors': errors,
            }

    def update_fee(self, garbage_fee_id, rental_area=None, unit_price=None,
                   min_amount=None, final_fee=None, status=None,
                   description=None, updated_by=None):
        with DBConnection() as conn:
            cursor = conn.cursor()

            cursor.execute("""
                SELECT GarbageFeeID, MerchantID, Year, FinalFee, ReceivableID, Status,
                       RentalArea, UnitPrice, MinAmount
                FROM GarbageFee WHERE GarbageFeeID = ?
            """, (garbage_fee_id,))
            row = cursor.fetchone()
            if not row:
                raise ValueError("垃圾费记录不存在")

            old_final_fee = float(row.FinalFee) if row.FinalFee else 0
            receivable_id = row.ReceivableID
            old_status = row.Status

            update_parts = []
            update_params = []

            if rental_area is not None:
                update_parts.append("RentalArea = ?")
                update_params.append(float(rental_area))

            if unit_price is not None:
                update_parts.append("UnitPrice = ?")
                update_params.append(float(unit_price))

            if min_amount is not None:
                update_parts.append("MinAmount = ?")
                update_params.append(float(min_amount))

            if rental_area is not None or unit_price is not None:
                new_area = float(rental_area) if rental_area is not None else float(row.RentalArea or 0)
                new_price = float(unit_price) if unit_price is not None else float(row.UnitPrice or 0)
                new_min = float(min_amount) if min_amount is not None else float(row.MinAmount or 0)
                calculated = round(new_price * new_area, 2)
                update_parts.append("CalculatedFee = ?")
                update_params.append(calculated)

                if final_fee is None:
                    final_fee = max(calculated, new_min)

            if final_fee is not None:
                update_parts.append("FinalFee = ?")
                update_params.append(float(final_fee))

            if status is not None:
                update_parts.append("Status = ?")
                update_params.append(status)

            if description is not None:
                update_parts.append("Description = ?")
                update_params.append(description)

            update_parts.append("UpdateBy = ?")
            update_params.append(updated_by)
            update_parts.append("UpdateTime = GETDATE()")

            update_params.append(garbage_fee_id)
            sql = f"UPDATE GarbageFee SET {', '.join(update_parts)} WHERE GarbageFeeID = ?"
            cursor.execute(sql, tuple(update_params))

            new_final_fee = float(final_fee) if final_fee is not None else old_final_fee

            if receivable_id:
                cursor.execute("""
                    SELECT ReceivableID, Status, PaidAmount, Amount
                    FROM Receivable WHERE ReceivableID = ? AND IsActive = 1
                """, (receivable_id,))
                rev_row = cursor.fetchone()

                if rev_row:
                    rev_status = rev_row.Status
                    rev_paid = float(rev_row.PaidAmount) if rev_row.PaidAmount else 0
                    rev_update_parts = []
                    rev_update_params = []

                    if abs(new_final_fee - old_final_fee) > 0.001:
                        if rev_status == '未付款':
                            rev_update_parts.append("Amount = ?")
                            rev_update_params.append(new_final_fee)
                            rev_update_parts.append("RemainingAmount = ?")
                            rev_update_params.append(new_final_fee)
                        elif rev_status == '部分付款':
                            new_remaining = new_final_fee - rev_paid
                            if new_remaining < 0:
                                raise ValueError(f"金额变更后剩余金额为负数（新总额{new_final_fee} - 已付{rev_paid}），请先处理收款记录")
                            rev_update_parts.append("Amount = ?")
                            rev_update_params.append(new_final_fee)
                            rev_update_parts.append("RemainingAmount = ?")
                            rev_update_params.append(new_remaining)
                            if new_remaining == 0:
                                rev_update_parts.append("Status = N'已付款'")

                    if status == '已收取' and rev_status != '已付款':
                        if 'Status' not in str(rev_update_parts):
                            rev_update_parts.append("Status = N'已付款'")
                        rev_update_parts.append("PaidAmount = Amount")

                    if rev_update_parts:
                        rev_update_params.append(receivable_id)
                        rev_sql = f"UPDATE Receivable SET {', '.join(rev_update_parts)} WHERE ReceivableID = ?"
                        cursor.execute(rev_sql, tuple(rev_update_params))

            conn.commit()

            logger.info(f"垃圾费记录更新成功: {garbage_fee_id}")
            return {'garbage_fee_id': garbage_fee_id}

    def delete_fee(self, garbage_fee_id):
        with DBConnection() as conn:
            cursor = conn.cursor()

            cursor.execute("SELECT GarbageFeeID, ReceivableID FROM GarbageFee WHERE GarbageFeeID = ?",
                           (garbage_fee_id,))
            row = cursor.fetchone()
            if not row:
                raise ValueError("垃圾费记录不存在")

            cursor.execute("DELETE FROM GarbageFee WHERE GarbageFeeID = ?", (garbage_fee_id,))

            conn.commit()
            logger.info(f"垃圾费记录删除成功: {garbage_fee_id}")
            return True

    def get_business_types(self):
        with DBConnection() as conn:
            cursor = conn.cursor()
            cursor.execute("""
                SELECT DictID, DictName, UnitPrice, MinAmount
                FROM Sys_Dictionary
                WHERE DictType = N'business_type' AND IsActive = 1
                ORDER BY SortOrder
            """)
            rows = cursor.fetchall()
            return [{'value': r.DictName, 'label': r.DictName,
                     'unit_price': float(r.UnitPrice) if r.UnitPrice else 0,
                     'min_amount': float(r.MinAmount) if r.MinAmount else 0}
                    for r in rows]

    def get_status_options(self):
        return [
            {'value': '待收取', 'label': '待收取'},
            {'value': '已收取', 'label': '已收取'},
        ]

    def export_fees(self, year=None, business_type=None, status=None, search=None):
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

        result = self.get_list(page=1, per_page=99999, year=year,
                               business_type=business_type, status=status, search=search)
        items = result['items']
        summary = result['summary']

        wb = Workbook()
        ws = wb.active
        ws.title = '垃圾费记录'

        header_font = Font(name='微软雅黑', bold=True, size=11, color='FFFFFF')
        header_fill = PatternFill(start_color='165DFF', end_color='165DFF', fill_type='solid')
        header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell_align_center = Alignment(horizontal='center', vertical='center')
        cell_align_right = Alignment(horizontal='right', vertical='center')
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin'),
        )
        summary_fill = PatternFill(start_color='E8F0FE', end_color='E8F0FE', fill_type='solid')
        summary_font = Font(name='微软雅黑', bold=True, size=11)

        title_font = Font(name='微软雅黑', bold=True, size=14)
        ws.merge_cells('A1:K1')
        ws['A1'] = f'{year or "全部年度"}垃圾费记录'
        ws['A1'].font = title_font
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 35

        headers = ['序号', '商户名称', '年度', '业态类型', '租赁面积(亩)',
                   '业态单价(元/亩/年)', '保底金额(元)', '计算金额(元)',
                   '最终金额(元)', '状态', '备注']
        header_row = 3
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border
        ws.row_dimensions[header_row].height = 30

        money_format = '#,##0.00'
        for idx, item in enumerate(items):
            row_num = header_row + 1 + idx
            data = [
                idx + 1,
                item['merchant_name'],
                item['year'],
                item['business_type'],
                item['rental_area'],
                item['unit_price'],
                item['min_amount'],
                item['calculated_fee'],
                item['final_fee'],
                item['status'],
                item['description'],
            ]
            aligns = [cell_align_center, cell_align_center, cell_align_center,
                      cell_align_center, cell_align_right, cell_align_right,
                      cell_align_right, cell_align_right, cell_align_right,
                      cell_align_center, Alignment(horizontal='left', vertical='center', wrap_text=True)]
            formats = [None, None, None, None, money_format, money_format,
                       money_format, money_format, money_format, None, None]
            for col_idx, (val, align, fmt) in enumerate(zip(data, aligns, formats), 1):
                cell = ws.cell(row=row_num, column=col_idx, value=val)
                cell.alignment = align
                cell.border = thin_border
                cell.font = Font(name='微软雅黑', size=10)
                if fmt:
                    cell.number_format = fmt

        summary_row = header_row + 1 + len(items)
        ws.merge_cells(f'A{summary_row}:D{summary_row}')
        summary_data = [
            ('合计', cell_align_center),
            ('', cell_align_center), ('', cell_align_center), ('', cell_align_center),
            (summary['total_area'], cell_align_right),
            ('', cell_align_right),
            ('', cell_align_right),
            ('', cell_align_right),
            (summary['total_fee'], cell_align_right),
            (f"{summary['total_count']}条", cell_align_center),
            ('', Alignment(horizontal='left', vertical='center')),
        ]
        for col_idx, (val, align) in enumerate(summary_data, 1):
            cell = ws.cell(row=summary_row, column=col_idx, value=val)
            cell.font = summary_font
            cell.fill = summary_fill
            cell.alignment = align
            cell.border = thin_border
            if col_idx in (5, 9):
                cell.number_format = money_format

        col_widths = [6, 18, 8, 12, 14, 18, 14, 14, 14, 8, 30]
        for i, width in enumerate(col_widths, 1):
            ws.column_dimensions[chr(64 + i)].width = width

        footer_row = summary_row + 1
        ws.merge_cells(f'A{footer_row}:K{footer_row}')
        ws.cell(row=footer_row, column=1,
                value=f'导出时间：{datetime.now().strftime("%Y-%m-%d %H:%M")}  共{len(items)}条记录')

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        date_str = datetime.now().strftime('%Y%m%d')
        filename = f'垃圾费记录_{date_str}.xlsx'
        return output, filename
