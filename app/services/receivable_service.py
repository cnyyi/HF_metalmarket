# -*- coding: utf-8 -*-
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from app.repositories.receivable_repo import ReceivableRepository
from utils.database import execute_query
from utils.format_utils import format_date, format_datetime


class ReceivableService:

    def __init__(self):
        self.repo = ReceivableRepository()

    def get_receivables(self, page=1, per_page=10, search=None, status=None, expense_type_id=None):
        rows, total_count, summaries = self.repo.get_list(page=page, per_page=per_page, search=search, status=status, expense_type_id=expense_type_id)
        total_pages = (total_count + per_page - 1) // per_page if total_count > 0 else 0

        result_list = []
        for row in rows:
            result_list.append({
                'receivable_id': row.ReceivableID,
                'merchant_id': row.MerchantID,
                'merchant_name': row.CustomerName or '',  # 统一使用 CASE 合并后的名称
                'customer_type': row.CustomerType or 'Merchant',
                'customer_id': row.CustomerID or row.MerchantID,
                'expense_type_id': row.ExpenseTypeID,
                'expense_type_name': row.ExpenseTypeName,
                'amount': float(row.Amount),
                'paid_amount': float(row.PaidAmount),
                'remaining_amount': float(row.RemainingAmount),
                'product_name': row.ProductName or '',
                'specification': row.Specification or '',
                'quantity': float(row.Quantity) if row.Quantity else None,
                'unit_id': row.UnitID,
                'unit_name': row.UnitName or '',
                'unit_price': float(row.UnitPrice) if row.UnitPrice else None,
                'due_date': format_date(row.DueDate),
                'status': row.Status,
                'description': row.Description or '',
                'create_time': format_datetime(row.CreateTime),
            })

        return {
            'items': result_list,
            'total_count': total_count,
            'total_pages': total_pages,
            'current_page': page,
            'summaries': summaries
        }

    def get_receivables_by_customer(self, page=1, per_page=10, search=None, status=None, expense_type_id=None):
        rows, total_count, sum_row = self.repo.get_list_by_customer(
            page=page, per_page=per_page, search=search, status=status, expense_type_id=expense_type_id
        )
        total_pages = (total_count + per_page - 1) // per_page if total_count > 0 else 0

        result_list = []
        for row in rows:
            result_list.append({
                'customer_type': row.CustomerType or 'Merchant',
                'customer_id': row.CustomerID,
                'customer_name': row.CustomerName or '',
                'total_amount': float(row.TotalAmount),
                'total_paid': float(row.TotalPaid),
                'total_remaining': float(row.TotalRemaining),
                'not_due_amount': float(row.NotDueAmount),
                'overdue_1to30': float(row.Overdue1to30),
                'overdue_31to60': float(row.Overdue31to60),
                'overdue_61to90': float(row.Overdue61to90),
                'overdue_over90': float(row.OverdueOver90),
                'record_count': row.RecordCount,
                'earliest_due_date': row.EarliestDueDate.strftime('%Y-%m-%d') if row.EarliestDueDate else '',
            })

        summary = {
            'total_amount': float(sum_row[0]),
            'total_paid': float(sum_row[1]),
            'total_remaining': float(sum_row[2]),
            'total_records': int(sum_row[3]),
            'not_due_amount': float(sum_row[4]),
            'overdue_1to30': float(sum_row[5]),
            'overdue_31to60': float(sum_row[6]),
            'overdue_61to90': float(sum_row[7]),
            'overdue_over90': float(sum_row[8]),
        } if sum_row else None

        return {
            'items': result_list,
            'total_count': total_count,
            'total_pages': total_pages,
            'current_page': page,
            'summary': summary,
        }

    def create_receivable(self, merchant_id=None, expense_type_id=None, amount=None, due_date=None,
                          description=None, reference_id=None, reference_type=None,
                          customer_type='Merchant', customer_id=None,
                          product_name=None, specification=None, quantity=None, unit_id=None, unit_price=None):
        """创建应收账款

        Args:
            merchant_id: 商户ID（兼容旧调用，Merchant 类型时必填）
            customer_type: 客户类型 'Merchant' 或 'Customer'
            customer_id: 客户ID（Merchant类型=MerchantID，Customer类型=CustomerID）
            product_name: 品名
            specification: 规格
            quantity: 数量
            unit_id: 单位ID（字典表unit_type）
            unit_price: 单价
        """
        if customer_type == 'Merchant':
            if not merchant_id and not customer_id:
                raise ValueError("请选择商户")
        else:
            if not customer_id:
                raise ValueError("请选择客户")

        if not expense_type_id:
            raise ValueError("请选择费用类型")
        if not amount or float(amount) <= 0:
            raise ValueError("应收金额必须大于0")
        if not due_date:
            raise ValueError("请选择到期日期")

        new_id = self.repo.create(
            merchant_id=int(merchant_id) if merchant_id else None,
            expense_type_id=int(expense_type_id),
            amount=float(amount),
            due_date=due_date,
            description=description,
            reference_id=reference_id,
            reference_type=reference_type,
            customer_type=customer_type,
            customer_id=int(customer_id) if customer_id else None,
            product_name=product_name,
            specification=specification,
            quantity=float(quantity) if quantity else None,
            unit_id=int(unit_id) if unit_id else None,
            unit_price=float(unit_price) if unit_price else None,
        )
        return new_id

    def pay(self, receivable_id, amount):
        receivable = self.repo.get_by_id(receivable_id)

        if not receivable:
            raise Exception("应收记录不存在")

        remaining = float(receivable.RemainingAmount)

        if amount <= 0:
            raise Exception("支付金额必须大于0")

        if amount > remaining:
            raise Exception("支付金额不能大于剩余金额")

        self.repo.update_payment(receivable_id, amount)

        new_remaining = remaining - amount

        if new_remaining == 0:
            self.repo.update_status(receivable_id, '已付款')
        else:
            self.repo.update_status(receivable_id, '部分付款')

        return True

    def soft_delete(self, receivable_id, deleted_by, delete_reason=None):
        """
        软删除应收账款

        业务规则：
        - 已付款/部分付款的应收禁止删除（只能通过冲销/退款处理）
        - 未付款的应收允许软删除，需记录删除原因
        - 已有关联收款记录/预收冲抵/押金转抵的应收禁止删除

        Args:
            receivable_id: 应收ID
            deleted_by: 删除操作人UserID
            delete_reason: 删除原因

        Returns:
            dict: {success, message}
        """
        # 1. 检查应收是否存在且有效
        receivable = self.repo.get_by_id(receivable_id, include_deleted=True)
        if not receivable:
            return {'success': False, 'message': '应收记录不存在'}

        if not receivable.IsActive:
            return {'success': False, 'message': '该应收已被删除'}

        # 2. 状态检查：已付款/部分付款不允许删除
        if receivable.Status in ('已付款', '部分付款'):
            return {'success': False, 'message': f'状态为"{receivable.Status}"的应收不允许删除，请通过冲销或退款处理'}

        # 3. 关联检查
        collection_count = self.repo.check_has_collection(receivable_id)
        if collection_count > 0:
            return {'success': False, 'message': f'该应收已有 {collection_count} 条收款记录，不允许删除'}

        prepay_count = self.repo.check_has_prepayment_apply(receivable_id)
        if prepay_count > 0:
            return {'success': False, 'message': f'该应收已有 {prepay_count} 条预收冲抵记录，不允许删除'}

        deposit_count = self.repo.check_has_deposit_transfer(receivable_id)
        if deposit_count > 0:
            return {'success': False, 'message': f'该应收已有 {deposit_count} 条押金转抵记录，不允许删除'}

        # 4. 执行软删除
        affected = self.repo.soft_delete(receivable_id, deleted_by, delete_reason)
        if affected > 0:
            return {'success': True, 'message': '删除成功'}
        else:
            return {'success': False, 'message': '删除失败，请重试'}

    def check_overdue(self):
        overdue_list = self.repo.list_overdue()

        for item in overdue_list:
            self.repo.update_status(item.ReceivableID, '逾期')

        return len(overdue_list)

    @staticmethod
    def get_expense_types(direction=None):
        query = """
            SELECT ExpenseTypeID, ExpenseTypeName, ExpenseTypeCode, ExpenseDirection, Description
            FROM ExpenseType
            WHERE IsActive = 1
        """
        params = []

        if direction:
            query += " AND ExpenseDirection = ?"
            params.append(direction)

        query += " ORDER BY ExpenseTypeID"

        results = execute_query(query, tuple(params) if params else None, fetch_type='all')

        return [{
            'expense_type_id': r.ExpenseTypeID,
            'expense_type_name': r.ExpenseTypeName,
            'expense_type_code': r.ExpenseTypeCode,
            'expense_direction': r.ExpenseDirection,
            'description': r.Description or ''
        } for r in results]

    @staticmethod
    def search_merchants(keyword):
        query = """
            SELECT MerchantID, MerchantName, ContactPerson, Phone
            FROM Merchant
            WHERE MerchantName LIKE ? OR ContactPerson LIKE ? OR Phone LIKE ?
            ORDER BY MerchantID
        """
        search_param = f'%{keyword}%'
        results = execute_query(query, (search_param, search_param, search_param), fetch_type='all')

        return [{
            'merchant_id': r.MerchantID,
            'merchant_name': r.MerchantName,
            'contact_person': r.ContactPerson or '',
            'phone': r.Phone or ''
        } for r in results]

    def export_receivables(self, search=None, status=None, expense_type_id=None):
        rows, total_count, summaries = self.repo.get_list(
            page=1, per_page=999999,
            search=search, status=status,
            expense_type_id=expense_type_id
        )

        items = []
        for row in rows:
            items.append({
                'receivable_id': row.ReceivableID,
                'customer_type': row.CustomerType or 'Merchant',
                'merchant_name': row.CustomerName or '',
                'expense_type_name': row.ExpenseTypeName or '',
                'amount': float(row.Amount),
                'paid_amount': float(row.PaidAmount),
                'remaining_amount': float(row.RemainingAmount),
                'product_name': row.ProductName or '',
                'specification': row.Specification or '',
                'quantity': float(row.Quantity) if row.Quantity else None,
                'unit_name': row.UnitName or '',
                'unit_price': float(row.UnitPrice) if row.UnitPrice else None,
                'due_date': format_date(row.DueDate),
                'status': row.Status,
                'description': row.Description or '',
            })

        wb = Workbook()
        ws = wb.active
        ws.title = '应收账款'

        header_font = Font(name='微软雅黑', bold=True, size=11, color='FFFFFF')
        header_fill = PatternFill(start_color='165DFF', end_color='165DFF', fill_type='solid')
        header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell_align_center = Alignment(horizontal='center', vertical='center')
        cell_align_right = Alignment(horizontal='right', vertical='center')
        cell_align_left = Alignment(horizontal='left', vertical='center', wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin'),
        )
        summary_fill = PatternFill(start_color='E8F0FE', end_color='E8F0FE', fill_type='solid')
        summary_font = Font(name='微软雅黑', bold=True, size=11)

        title_font = Font(name='微软雅黑', bold=True, size=14)
        ws.merge_cells('A1:L1')
        ws['A1'] = '应收账款'
        ws['A1'].font = title_font
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 35

        filter_parts = []
        if search:
            filter_parts.append(f'搜索：{search}')
        if status:
            filter_parts.append(f'状态：{status}')
        if expense_type_id:
            filter_parts.append(f'费用类型ID：{expense_type_id}')
        filter_text = '  |  '.join(filter_parts) if filter_parts else '全部记录'
        ws.merge_cells('A2:L2')
        ws['A2'] = filter_text
        ws['A2'].font = Font(name='微软雅黑', size=9, color='666666')
        ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[2].height = 22

        headers = ['序号', '客户类型', '名称', '费用类型', '应收金额(元)', '已收金额(元)',
                   '未收金额(元)', '到期日期', '状态', '品名', '规格', '备注']
        header_row = 3
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border
        ws.row_dimensions[header_row].height = 30

        money_format = '#,##0.00'
        for idx, item in enumerate(items):
            row_num = header_row + 1 + idx
            data = [
                idx + 1,
                '往来客户' if item['customer_type'] == 'Customer' else '内部商户',
                item['merchant_name'],
                item['expense_type_name'],
                item['amount'],
                item['paid_amount'],
                item['remaining_amount'],
                item['due_date'],
                item['status'],
                item['product_name'],
                item['specification'],
                item['description'],
            ]
            aligns = [
                cell_align_center, cell_align_center, cell_align_left,
                cell_align_center, cell_align_right, cell_align_right,
                cell_align_right, cell_align_center, cell_align_center,
                cell_align_left, cell_align_left, cell_align_left,
            ]
            formats = [
                None, None, None, None,
                money_format, money_format, money_format,
                None, None, None, None, None,
            ]
            for col_idx, (val, align, fmt) in enumerate(zip(data, aligns, formats), 1):
                cell = ws.cell(row=row_num, column=col_idx, value=val)
                cell.alignment = align
                cell.border = thin_border
                cell.font = Font(name='微软雅黑', size=10)
                if fmt:
                    cell.number_format = fmt

        summary_row = header_row + 1 + len(items)
        ws.merge_cells(f'A{summary_row}:D{summary_row}')
        # 合并区域只写左上角，其余列为 MergedCell 不可赋值
        summary_cell_a = ws.cell(row=summary_row, column=1,
                                 value=f'合计（共{total_count}条）')
        summary_cell_a.font = summary_font
        summary_cell_a.fill = summary_fill
        summary_cell_a.alignment = cell_align_center
        summary_cell_a.border = thin_border
        # B/C/D 列设置样式（不赋值，仅边框和填充）
        for col_idx in (2, 3, 4):
            cell = ws.cell(row=summary_row, column=col_idx)
            cell.font = summary_font
            cell.fill = summary_fill
            cell.alignment = cell_align_center
            cell.border = thin_border
        # E-L 列正常赋值
        summary_right = [
            (5, summaries['total_amount'], cell_align_right),
            (6, summaries['total_paid'], cell_align_right),
            (7, summaries['total_remaining'], cell_align_right),
        ]
        for col_idx, val, align in summary_right:
            cell = ws.cell(row=summary_row, column=col_idx, value=val)
            cell.font = summary_font
            cell.fill = summary_fill
            cell.alignment = align
            cell.border = thin_border
            cell.number_format = money_format
        for col_idx in (8, 9, 10, 11, 12):
            cell = ws.cell(row=summary_row, column=col_idx, value='')
            cell.font = summary_font
            cell.fill = summary_fill
            cell.alignment = cell_align_center
            cell.border = thin_border

        col_widths = [6, 10, 18, 12, 14, 14, 14, 12, 8, 10, 10, 30]
        for i, width in enumerate(col_widths, 1):
            ws.column_dimensions[chr(64 + i)].width = width

        from datetime import datetime as dt
        footer_row = summary_row + 1
        ws.merge_cells(f'A{footer_row}:L{footer_row}')
        ws.cell(row=footer_row, column=1,
                value=f'导出时间：{dt.now().strftime("%Y-%m-%d %H:%M")}  共{len(items)}条记录')

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        date_str = dt.now().strftime('%Y%m%d')
        filename = f'应收账款_{date_str}.xlsx'
        return output, filename

    # ========== Agent 查询方法 ==========

    def get_receivable_summary(self, group_by_merchant=False, merchant_id=None, source='admin'):
        with DBConnection() as conn:
            cursor = conn.cursor()
            merchant_filter = ''
            params = []
            if source == 'wx' and merchant_id:
                merchant_filter = "AND r.MerchantID = ?"
                params = [merchant_id]
            if group_by_merchant:
                cursor.execute(f"""
                    SELECT m.MerchantName,
                           SUM(r.Amount) AS total_amount,
                           SUM(r.PaidAmount) AS paid_amount,
                           SUM(r.RemainingAmount) AS remaining_amount,
                           COUNT(*) AS count
                    FROM Receivable r
                    INNER JOIN Merchant m ON r.MerchantID = m.MerchantID
                    WHERE r.IsActive = 1 {merchant_filter}
                    GROUP BY m.MerchantName
                    ORDER BY remaining_amount DESC
                """, params)
                rows = cursor.fetchall()
                return [{'merchant_name': row.MerchantName, 'count': row.count,
                         'total_amount': round(float(row.total_amount or 0), 2),
                         'paid_amount': round(float(row.paid_amount or 0), 2),
                         'remaining_amount': round(float(row.remaining_amount or 0), 2)} for row in rows]
            else:
                cursor.execute(f"""
                    SELECT COUNT(*) AS total_count,
                           SUM(r.Amount) AS total_amount,
                           SUM(r.PaidAmount) AS paid_amount,
                           SUM(r.RemainingAmount) AS remaining_amount
                    FROM Receivable r
                    WHERE r.IsActive = 1 {merchant_filter}
                """, params)
                row = cursor.fetchone()
                return {'total_count': row.total_count,
                        'total_amount': round(float(row.total_amount or 0), 2),
                        'paid_amount': round(float(row.paid_amount or 0), 2),
                        'remaining_amount': round(float(row.remaining_amount or 0), 2)}

    def get_overdue_receivables(self, merchant_id=None, source='admin'):
        with DBConnection() as conn:
            cursor = conn.cursor()
            merchant_filter = ''
            params = []
            if source == 'wx' and merchant_id:
                merchant_filter = "AND r.MerchantID = ?"
                params = [merchant_id]
            cursor.execute(f"""
                SELECT TOP 500 r.ReceivableID, m.MerchantName, r.Amount, r.PaidAmount,
                       r.RemainingAmount, r.DueDate, r.Status,
                       sd.DictName AS expense_type_name
                FROM Receivable r
                INNER JOIN Merchant m ON r.MerchantID = m.MerchantID
                LEFT JOIN Sys_Dictionary sd ON r.ExpenseTypeID = sd.DictID
                WHERE r.IsActive = 1
                  AND r.DueDate < CAST(GETDATE() AS DATE)
                  AND r.Status IN (N'未付款', N'部分付款')
                  AND r.RemainingAmount > 0
                  {merchant_filter}
                ORDER BY r.DueDate
            """, params)
            rows = cursor.fetchall()
            return [{'receivable_id': row.ReceivableID, 'merchant_name': row.MerchantName,
                     'amount': round(float(row.Amount or 0), 2),
                     'paid_amount': round(float(row.PaidAmount or 0), 2),
                     'remaining_amount': round(float(row.RemainingAmount or 0), 2),
                     'due_date': row.DueDate.strftime('%Y-%m-%d') if row.DueDate else '',
                     'status': row.Status, 'expense_type': row.expense_type_name} for row in rows]
