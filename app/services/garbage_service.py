# -*- coding: utf-8 -*-
"""
垃圾清运管理服务层
负责垃圾清运记录的创建、查询、编辑、删除等业务逻辑
垃圾清运 → 市场方应付供应商（联动 Payable）
"""
import logging
from datetime import datetime, date
from utils.database import DBConnection

logger = logging.getLogger(__name__)

# 处置费字典 DictID（expense_item_expend）
GARBAGE_EXPENSE_TYPE_ID = 1075  # disposal_fee 处置费
GARBAGE_EXPENSE_TYPE_NAME = '处置费'


class GarbageService:
    """垃圾清运管理服务"""

    # ========== 列表查询 ==========

    def get_collections(self, page=1, per_page=10, search=None, date_from=None,
                        date_to=None, vendor_id=None, show_all=False):
        """
        获取垃圾清运记录分页列表
        筛选条件：关键词搜索 / 日期范围 / 供应商
        show_all=True 时返回全部数据（不分页）
        """
        with DBConnection() as conn:
            cursor = conn.cursor()

            base_query = """
                SELECT gc.CollectionID, gc.CollectionDate, gc.GarbageType, gc.Amount,
                       gc.Unit, gc.UnitPrice, gc.TotalAmount, gc.Status, gc.Description,
                       gc.PayableID, gc.CreateBy, gc.CreateTime, gc.UpdateBy, gc.UpdateTime,
                       c.CustomerName AS VendorName,
                       u.RealName AS CreateUserName
                FROM GarbageCollection gc
                LEFT JOIN Customer c ON gc.CustomerID = c.CustomerID
                LEFT JOIN [User] u ON gc.CreateBy = u.UserID
            """
            count_query = """
                SELECT COUNT(*)
                FROM GarbageCollection gc
                LEFT JOIN Customer c ON gc.CustomerID = c.CustomerID
            """
            summary_query = """
                SELECT ISNULL(COUNT(*), 0),
                       ISNULL(SUM(gc.Amount), 0),
                       ISNULL(SUM(gc.TotalAmount), 0)
                FROM GarbageCollection gc
                LEFT JOIN Customer c ON gc.CustomerID = c.CustomerID
            """

            conditions = []
            params = []

            if search:
                conditions.append("(c.CustomerName LIKE ? OR gc.GarbageType LIKE ? OR gc.Description LIKE ?)")
                p = f'%{search}%'
                params.extend([p, p, p])

            if date_from:
                conditions.append("gc.CollectionDate >= ?")
                params.append(date_from)

            if date_to:
                conditions.append("gc.CollectionDate <= ?")
                params.append(date_to)

            if vendor_id:
                conditions.append("gc.CustomerID = ?")
                params.append(int(vendor_id))

            where_clause = ""
            if conditions:
                where_clause = " WHERE " + " AND ".join(conditions)
                base_query += where_clause
                count_query += where_clause
                summary_query += where_clause

            cursor.execute(count_query, tuple(params))
            total_count = cursor.fetchone()[0]

            cursor.execute(summary_query, tuple(params))
            summary_row = cursor.fetchone()
            summary = {
                'total_count': summary_row[0],
                'total_amount': float(summary_row[2]),
                'total_quantity': float(summary_row[1]),
            }

            if show_all:
                base_query += " ORDER BY gc.CollectionDate DESC, gc.CollectionID DESC"
            else:
                offset = (page - 1) * per_page
                base_query += " ORDER BY gc.CollectionDate DESC, gc.CollectionID DESC OFFSET ? ROWS FETCH NEXT ? ROWS ONLY"
                params.extend([offset, per_page])

            cursor.execute(base_query, tuple(params))
            rows = cursor.fetchall()

            items = []
            for row in rows:
                coll_date = row.CollectionDate
                if coll_date:
                    coll_date_str = coll_date.strftime('%Y-%m-%d') if hasattr(coll_date, 'strftime') else str(coll_date)[:10]
                else:
                    coll_date_str = ''
                create_time = row.CreateTime
                if create_time:
                    create_time_str = create_time.strftime('%Y-%m-%d %H:%M') if hasattr(create_time, 'strftime') else str(create_time)[:16]
                else:
                    create_time_str = ''
                items.append({
                    'collection_id': row.CollectionID,
                    'collection_date': coll_date_str,
                    'vendor_name': row.VendorName or '',
                    'garbage_type': row.GarbageType or '',
                    'amount': float(row.Amount) if row.Amount else 0,
                    'unit': row.Unit or '',
                    'unit_price': float(row.UnitPrice) if row.UnitPrice else 0,
                    'total_amount': float(row.TotalAmount) if row.TotalAmount else 0,
                    'status': row.Status or '',
                    'description': row.Description or '',
                    'payable_id': row.PayableID,
                    'create_user_name': row.CreateUserName or '',
                    'create_time': create_time_str,
                })

            total_pages = (total_count + per_page - 1) // per_page if total_count > 0 else 1

            result = {
                'items': items,
                'total': total_count,
                'page': page,
                'per_page': per_page,
                'total_pages': total_pages,
                'summary': summary,
            }

            if show_all:
                result['page'] = 1
                result['per_page'] = total_count
                result['total_pages'] = 1

            return result

    # ========== 详情查询 ==========

    def get_collection_detail(self, collection_id):
        """获取垃圾清运记录详情"""
        with DBConnection() as conn:
            cursor = conn.cursor()

            cursor.execute("""
                SELECT gc.CollectionID, gc.CollectionDate, gc.CustomerID, gc.GarbageType,
                       gc.Amount, gc.Unit, gc.UnitPrice, gc.TotalAmount,
                       gc.Status, gc.Description, gc.PayableID,
                       gc.CreateBy, gc.CreateTime, gc.UpdateBy, gc.UpdateTime,
                       c.CustomerName AS VendorName,
                       u.RealName AS CreateUserName
                FROM GarbageCollection gc
                LEFT JOIN Customer c ON gc.CustomerID = c.CustomerID
                LEFT JOIN [User] u ON gc.CreateBy = u.UserID
                WHERE gc.CollectionID = ?
            """, (collection_id,))
            row = cursor.fetchone()
            if not row:
                return None

            # CollectionDate 可能是 str（pyodbc 对 DATE 的处理）或 date/datetime
            coll_date = row.CollectionDate
            if coll_date:
                coll_date_str = coll_date.strftime('%Y-%m-%d') if hasattr(coll_date, 'strftime') else str(coll_date)[:10]
            else:
                coll_date_str = ''
            create_time = row.CreateTime
            if create_time:
                create_time_str = create_time.strftime('%Y-%m-%d %H:%M') if hasattr(create_time, 'strftime') else str(create_time)[:16]
            else:
                create_time_str = ''
            return {
                'collection_id': row.CollectionID,
                'collection_date': coll_date_str,
                'customer_id': row.CustomerID,
                'vendor_name': row.VendorName or '',
                'garbage_type': row.GarbageType or '',
                'amount': float(row.Amount) if row.Amount else 0,
                'unit': row.Unit or '',
                'unit_price': float(row.UnitPrice) if row.UnitPrice else 0,
                'total_amount': float(row.TotalAmount) if row.TotalAmount else 0,
                'status': row.Status or '',
                'description': row.Description or '',
                'payable_id': row.PayableID,
                'create_user_name': row.CreateUserName or '',
                'create_time': create_time_str,
            }

    # ========== 创建垃圾清运记录（联动 Payable） ==========

    def create_collection(self, collection_date, customer_id, garbage_type,
                         amount, unit, unit_price, total_amount, description=None, created_by=None):
        """
        创建垃圾清运记录，同时在 Payable 表创建对应的应付账款

        Args:
            collection_date: 清运日期 (str YYYY-MM-DD)
            customer_id: 供应商 CustomerID
            garbage_type: 垃圾类型
            amount: 数量
            unit: 单位
            unit_price: 单价
            total_amount: 总金额
            description: 备注
            created_by: 创建人 UserID

        Returns:
            dict: {'collection_id': int, 'payable_id': int}
        """
        # 参数校验
        if not collection_date:
            raise ValueError("请选择清运日期")
        if not customer_id:
            raise ValueError("请选择供应商")
        if not garbage_type:
            raise ValueError("请选择垃圾类型")
        if not amount or float(amount) <= 0:
            raise ValueError("数量必须大于0")
        if not unit:
            raise ValueError("请选择单位")
        if not unit_price or float(unit_price) <= 0:
            raise ValueError("单价必须大于0")
        if not total_amount or float(total_amount) <= 0:
            raise ValueError("总金额必须大于0")

        with DBConnection() as conn:
            cursor = conn.cursor()

            # 查询供应商名称
            cursor.execute("SELECT CustomerName FROM Customer WHERE CustomerID = ?", (customer_id,))
            row = cursor.fetchone()
            if not row:
                raise ValueError("供应商不存在，请重新选择")
            vendor_name = row.CustomerName

            # Step 1: 插入垃圾清运记录
            cursor.execute("""
                INSERT INTO GarbageCollection (CollectionDate, CustomerID, GarbageType,
                                               Amount, Unit, UnitPrice, TotalAmount,
                                               Status, Description, CreateBy, CreateTime)
                OUTPUT INSERTED.CollectionID
                VALUES (?, ?, ?, ?, ?, ?, ?, N'待结算', ?, ?, GETDATE())
            """, (
                collection_date, customer_id, garbage_type,
                amount, unit, unit_price, total_amount,
                description, created_by
            ))
            gc_row = cursor.fetchone()
            collection_id = gc_row[0]

            # Step 2: 联动创建 Payable（应付账款）
            # DueDate = CollectionDate（当天到期）
            cursor.execute("""
                INSERT INTO Payable (VendorName, ExpenseTypeID, Amount, DueDate,
                                     Status, PaidAmount, RemainingAmount, Description,
                                     CustomerType, CustomerID,
                                     ReferenceType, ReferenceID)
                OUTPUT INSERTED.PayableID
                VALUES (?, ?, ?, ?, N'未付款', 0, ?, ?, N'服务商', ?, N'garbage_collection', ?)
            """, (
                vendor_name,
                GARBAGE_EXPENSE_TYPE_ID,
                total_amount,
                collection_date,          # DueDate = 当天到期
                total_amount,
                f'垃圾清运：{garbage_type}，数量{amount}{unit}',
                customer_id,
                collection_id
            ))
            payable_row = cursor.fetchone()
            payable_id = payable_row[0]

            # Step 3: 回写 PayableID 到 GarbageCollection
            cursor.execute("""
                UPDATE GarbageCollection SET PayableID = ? WHERE CollectionID = ?
            """, (payable_id, collection_id))

            conn.commit()

            logger.info(f"垃圾清运记录创建成功: CollectionID={collection_id}, PayableID={payable_id}")
            return {
                'collection_id': collection_id,
                'payable_id': payable_id,
            }

    # ========== 更新垃圾清运记录 ==========

    def update_collection(self, collection_id, collection_date, customer_id, garbage_type,
                          amount, unit, unit_price, total_amount, description=None,
                          status=None, updated_by=None):
        """
        更新垃圾清运记录，联动更新关联的 Payable 数据

        联动规则：
        - 供应商变更 → 更新 Payable VendorName/CustomerID
        - 金额变更 → 更新 Payable Amount/RemainingAmount（仅未付款状态）
        - 日期变更 → 更新 Payable DueDate
        - 状态→已结算 → 更新 Payable 状态为已付款
        """
        if not collection_date:
            raise ValueError("请选择清运日期")
        if not customer_id:
            raise ValueError("请选择供应商")
        if not garbage_type:
            raise ValueError("请选择垃圾类型")
        if not amount or float(amount) <= 0:
            raise ValueError("数量必须大于0")
        if not unit:
            raise ValueError("请选择单位")
        if not unit_price or float(unit_price) <= 0:
            raise ValueError("单价必须大于0")
        if not total_amount or float(total_amount) <= 0:
            raise ValueError("总金额必须大于0")
        if not status:
            raise ValueError("请选择状态")

        with DBConnection() as conn:
            cursor = conn.cursor()

            # 查询现有记录及关联 PayableID + 旧供应商信息
            cursor.execute("""
                SELECT gc.CollectionID, gc.PayableID, gc.TotalAmount, gc.CustomerID,
                       c.CustomerName
                FROM GarbageCollection gc
                LEFT JOIN Customer c ON gc.CustomerID = c.CustomerID
                WHERE gc.CollectionID = ?
            """, (collection_id,))
            row = cursor.fetchone()
            if not row:
                raise ValueError("垃圾清运记录不存在")

            payable_id = row.PayableID
            old_total = float(row.TotalAmount) if row.TotalAmount else 0
            old_customer_id = row.CustomerID

            # 查询新供应商名称
            cursor.execute("SELECT CustomerName FROM Customer WHERE CustomerID = ?", (customer_id,))
            vendor_row = cursor.fetchone()
            new_vendor_name = vendor_row.CustomerName if vendor_row else ''

            # 更新垃圾清运记录
            cursor.execute("""
                UPDATE GarbageCollection SET
                    CollectionDate = ?,
                    CustomerID = ?,
                    GarbageType = ?,
                    Amount = ?,
                    Unit = ?,
                    UnitPrice = ?,
                    TotalAmount = ?,
                    Description = ?,
                    Status = ?,
                    UpdateBy = ?,
                    UpdateTime = GETDATE()
                WHERE CollectionID = ?
            """, (
                collection_date, customer_id, garbage_type,
                amount, unit, unit_price, total_amount,
                description, status, updated_by, collection_id
            ))

            # 联动更新关联的 Payable
            if payable_id:
                # 查询 Payable 当前状态
                cursor.execute("""
                    SELECT PayableID, Status, PaidAmount, Amount, CustomerID
                    FROM Payable WHERE PayableID = ? AND IsActive = 1
                """, (payable_id,))
                pay_row = cursor.fetchone()

                if pay_row:
                    pay_status = pay_row.Status
                    pay_paid = float(pay_row.PaidAmount) if pay_row.PaidAmount else 0
                    new_total = float(total_amount)

                    # 构建动态更新字段
                    update_parts = []
                    update_params = []

                    # 供应商变更 → 更新 VendorName + CustomerID
                    if int(customer_id) != (old_customer_id or 0):
                        update_parts.append("VendorName = ?")
                        update_params.append(new_vendor_name)
                        update_parts.append("CustomerID = ?")
                        update_params.append(customer_id)

                    # 日期变更 → 更新 DueDate
                    update_parts.append("DueDate = ?")
                    update_params.append(collection_date)

                    # 描述更新
                    update_parts.append("Description = ?")
                    update_params.append(f'垃圾清运：{garbage_type}，数量{amount}{unit}')

                    # 金额变更处理
                    if abs(new_total - old_total) > 0.001:
                        if pay_status == '未付款':
                            # 未付款：直接更新金额和剩余金额
                            update_parts.append("Amount = ?")
                            update_params.append(new_total)
                            update_parts.append("RemainingAmount = ?")
                            update_params.append(new_total)
                        elif pay_status == '部分付款':
                            # 部分付款：更新总额，剩余金额 = 新总额 - 已付金额
                            new_remaining = new_total - pay_paid
                            if new_remaining < 0:
                                raise ValueError(f"金额变更后剩余金额为负数（新总额{new_total} - 已付{pay_paid}），请先处理付款记录")
                            update_parts.append("Amount = ?")
                            update_params.append(new_total)
                            update_parts.append("RemainingAmount = ?")
                            update_params.append(new_remaining)
                            # 如果剩余金额为0，自动变为已付款
                            if new_remaining == 0:
                                update_parts.append("Status = N'已付款'")
                        # 已付款状态不允许改金额（有付款记录锁定）

                    # 状态→已结算 → 更新 Payable 状态为已付款
                    if status == '已结算' and pay_status != '已付款':
                        if 'Status' not in str(update_parts):
                            update_parts.append("Status = N'已付款'")
                        update_parts.append("PaidAmount = Amount")

                    if update_parts:
                        sql = f"UPDATE Payable SET {', '.join(update_parts)} WHERE PayableID = ?"
                        update_params.append(payable_id)
                        cursor.execute(sql, tuple(update_params))

            conn.commit()

            logger.info(f"垃圾清运记录更新成功: {collection_id}, Payable联动更新: {payable_id}")
            return {
                'collection_id': collection_id,
            }

    # ========== 删除垃圾清运记录 ==========

    def delete_collection(self, collection_id):
        """
        删除垃圾清运记录（需先确保 Payable 已无未核销余额或单独处理）
        """
        with DBConnection() as conn:
            cursor = conn.cursor()

            # 查询关联 Payable
            cursor.execute("SELECT PayableID FROM GarbageCollection WHERE CollectionID = ?", (collection_id,))
            row = cursor.fetchone()
            if not row:
                raise ValueError("垃圾清运记录不存在")

            # 删除垃圾清运记录
            cursor.execute("DELETE FROM GarbageCollection WHERE CollectionID = ?", (collection_id,))

            # Payable 保留（核销记录），由应付管理模块处理
            conn.commit()

            logger.info(f"垃圾清运记录删除成功: {collection_id}")
            return True

    # ========== 字典接口 ==========

    def get_vendors(self):
        """
        获取供应商列表（从 Customer 表筛选"服务商"类型）
        """
        with DBConnection() as conn:
            cursor = conn.cursor()

            cursor.execute("""
                SELECT CustomerID, CustomerName, CustomerType
                FROM Customer
                WHERE CustomerType = N'服务商'
                  AND Status = N'正常'
                ORDER BY CustomerName
            """)
            rows = cursor.fetchall()

            vendors = []
            for row in rows:
                vendors.append({
                    'vendor_id': row.CustomerID,
                    'vendor_name': row.CustomerName,
                    'vendor_type': row.CustomerType or '',
                })

            return vendors

    def get_garbage_types(self):
        """
        获取垃圾类型列表（从字典表获取）
        """
        with DBConnection() as conn:
            cursor = conn.cursor()

            # 优先从字典表获取，否则返回固定列表
            try:
                cursor.execute("""
                    SELECT DictID, DictName
                    FROM Sys_Dictionary
                    WHERE DictType = 'garbage_type' AND IsActive = 1
                    ORDER BY SortOrder
                """)
                rows = cursor.fetchall()
                if rows:
                    return [{'value': str(r.DictID), 'label': r.DictName} for r in rows]
            except Exception:
                pass

            # fallback：返回固定列表
            return [
                {'value': '生活垃圾', 'label': '生活垃圾'},
                {'value': '建筑垃圾', 'label': '建筑垃圾'},
                {'value': '工业垃圾', 'label': '工业垃圾'},
                {'value': '其他垃圾', 'label': '其他垃圾'},
            ]

    def get_status_options(self):
        """获取状态选项"""
        return [
            {'value': '待结算', 'label': '待结算'},
            {'value': '已结算', 'label': '已结算'},
        ]

    # ========== 统计 ==========

    def get_summary(self, date_from=None, date_to=None, vendor_id=None):
        """获取垃圾清运汇总统计"""
        with DBConnection() as conn:
            cursor = conn.cursor()

            query = "SELECT ISNULL(SUM(TotalAmount), 0) FROM GarbageCollection"
            conditions = []
            params = []

            if date_from:
                conditions.append("CollectionDate >= ?")
                params.append(date_from)
            if date_to:
                conditions.append("CollectionDate <= ?")
                params.append(date_to)
            if vendor_id:
                conditions.append("CustomerID = ?")
                params.append(vendor_id)

            if conditions:
                query += " WHERE " + " AND ".join(conditions)

            cursor.execute(query, tuple(params))
            row = cursor.fetchone()
            return float(row[0])

    # ========== 导出 Excel ==========

    def export_collections(self, search=None, date_from=None, date_to=None, vendor_id=None):
        """
        导出垃圾清运记录为 Excel 文件
        返回 (BytesIO, filename) 元组
        """
        from io import BytesIO
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, numbers

        # 复用 get_collections 获取全部数据
        result = self.get_collections(
            page=1, per_page=99999,
            search=search, date_from=date_from, date_to=date_to,
            vendor_id=vendor_id, show_all=True,
        )
        items = result['items']
        summary = result['summary']

        wb = Workbook()
        ws = wb.active
        ws.title = '垃圾清运记录'

        # 样式定义
        header_font = Font(name='微软雅黑', bold=True, size=11, color='FFFFFF')
        header_fill = PatternFill(start_color='165DFF', end_color='165DFF', fill_type='solid')
        header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell_align_center = Alignment(horizontal='center', vertical='center')
        cell_align_right = Alignment(horizontal='right', vertical='center')
        cell_align_left = Alignment(horizontal='left', vertical='center', wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin'),
        )
        summary_fill = PatternFill(start_color='E8F0FE', end_color='E8F0FE', fill_type='solid')
        summary_font = Font(name='微软雅黑', bold=True, size=11)

        # 标题行
        title_font = Font(name='微软雅黑', bold=True, size=14)
        ws.merge_cells('A1:J1')
        ws['A1'] = '垃圾清运记录'
        ws['A1'].font = title_font
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 35

        # 筛选条件行
        filter_parts = []
        if date_from or date_to:
            filter_parts.append(f'日期：{date_from or "起始"} ~ {date_to or "至今"}')
        if vendor_id:
            filter_parts.append(f'供应商ID：{vendor_id}')
        if search:
            filter_parts.append(f'搜索：{search}')
        filter_text = '  |  '.join(filter_parts) if filter_parts else '全部记录'
        ws.merge_cells('A2:J2')
        ws['A2'] = filter_text
        ws['A2'].font = Font(name='微软雅黑', size=9, color='666666')
        ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[2].height = 22

        # 表头
        headers = ['序号', '清运日期', '供应商', '垃圾类型', '数量', '单位', '单价(元)', '总金额(元)', '状态', '备注']
        header_row = 3
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border
        ws.row_dimensions[header_row].height = 30

        # 数据行
        money_format = '#,##0.00'
        quantity_format = '#,##0.00'
        for idx, item in enumerate(items):
            row_num = header_row + 1 + idx
            data = [
                idx + 1,
                item['collection_date'],
                item['vendor_name'],
                item['garbage_type'],
                item['amount'],
                item['unit'],
                item['unit_price'],
                item['total_amount'],
                item['status'],
                item['description'],
            ]
            aligns = [
                cell_align_center, cell_align_center, cell_align_center,
                cell_align_center, cell_align_right, cell_align_center,
                cell_align_right, cell_align_right, cell_align_center,
                cell_align_left,
            ]
            formats = [
                None, None, None, None,
                quantity_format, None,
                money_format, money_format, None, None,
            ]
            for col_idx, (val, align, fmt) in enumerate(zip(data, aligns, formats), 1):
                cell = ws.cell(row=row_num, column=col_idx, value=val)
                cell.alignment = align
                cell.border = thin_border
                cell.font = Font(name='微软雅黑', size=10)
                if fmt:
                    cell.number_format = fmt

        # 合计行
        summary_row = header_row + 1 + len(items)
        ws.merge_cells(f'A{summary_row}:D{summary_row}')
        summary_data = [
            ('合计', cell_align_center),
            ('', cell_align_center), ('', cell_align_center), ('', cell_align_center),
            (summary['total_quantity'], cell_align_right),
            ('', cell_align_center),
            ('', cell_align_right),
            (summary['total_amount'], cell_align_right),
            (f"{summary['total_count']}条", cell_align_center),
            ('', cell_align_left),
        ]
        for col_idx, (val, align) in enumerate(summary_data, 1):
            cell = ws.cell(row=summary_row, column=col_idx, value=val)
            cell.font = summary_font
            cell.fill = summary_fill
            cell.alignment = align
            cell.border = thin_border
            if col_idx == 5:
                cell.number_format = quantity_format
            elif col_idx == 8:
                cell.number_format = money_format

        # 列宽
        col_widths = [6, 12, 18, 12, 10, 8, 12, 14, 8, 30]
        for i, width in enumerate(col_widths, 1):
            ws.column_dimensions[chr(64 + i)].width = width

        # 导出时间戳（页脚）
        from datetime import datetime
        footer_row = summary_row + 1
        ws.merge_cells(f'A{footer_row}:J{footer_row}')
        ws.cell(row=footer_row, column=1,
                value=f'导出时间：{datetime.now().strftime("%Y-%m-%d %H:%M")}  共{len(items)}条记录')

        # 保存到 BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        date_str = datetime.now().strftime('%Y%m%d')
        filename = f'垃圾清运记录_{date_str}.xlsx'
        return output, filename

    def get_garbage_fees(self, year=None, merchant_id=None, source='admin'):
        year_filter = ''
        params = []
        if year:
            year_filter = "AND gf.Year = ?"
            params = [year]
        merchant_filter = ''
        if source == 'wx' and merchant_id:
            merchant_filter = "AND gf.MerchantID = ?"
            params.append(merchant_id)
        with DBConnection() as conn:
            cursor = conn.cursor()
            cursor.execute(f"""
                SELECT TOP 500 gf.GarbageFeeID, m.MerchantName, gf.Year,
                       gf.BusinessType, gf.Area, gf.FeeAmount
                FROM GarbageFee gf
                INNER JOIN Merchant m ON gf.MerchantID = m.MerchantID
                WHERE 1=1 {year_filter} {merchant_filter}
                ORDER BY gf.Year DESC, gf.FeeAmount DESC
            """, params)
            rows = cursor.fetchall()
            return [{'garbage_fee_id': row.GarbageFeeID, 'merchant_name': row.MerchantName,
                     'year': row.Year, 'business_type': row.BusinessType,
                     'area': round(float(row.Area or 0), 2),
                     'fee_amount': round(float(row.FeeAmount or 0), 2)} for row in rows]
